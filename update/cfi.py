#!/usr/bin/env python3

# update/cfi.py - script to download CFI code list from the SIX group
#
# Copyright (C) 2022-2024 Arthur de Jong
#
# This library is free software; you can redistribute it and/or
# modify it under the terms of the GNU Lesser General Public
# License as published by the Free Software Foundation; either
# version 2.1 of the License, or (at your option) any later version.
#
# This library is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the GNU
# Lesser General Public License for more details.
#
# You should have received a copy of the GNU Lesser General Public
# License along with this library; if not, write to the Free Software
# Foundation, Inc., 51 Franklin Street, Fifth Floor, Boston, MA
# 02110-1301 USA

"""This script downloads the list of CFI codes as published by the SIX group."""

import io
import re

import lxml.html
import openpyxl
import requests


# the location of the Statistical Classification file
download_url = 'https://www.six-group.com/en/products-services/financial-information/data-standards.html'


def normalise(value):
    """Clean and minimise attribute names and values."""
    return re.sub(r' *[(\[\n].*', '', value, re.MULTILINE).strip()


def get_categories(sheet):
    """Get the list of top-level CFI categories."""
    for row in sheet.iter_rows():
        if row[0].value and len(row[0].value) == 1 and row[1].value:
            yield (row[0].value, row[1].value)


def get_attributes(sheet):
    """Get the list of characters and attributes from the group-specific sheet."""
    attribute = None
    value_list = []
    values = None
    for row in sheet.iter_rows():
        if row[0].value and not row[1].value and row[2].value:
            attribute = normalise(row[2].value)
            values = []
            value_list.append((attribute, values))
        elif attribute and row[1].value and row[2].value:
            values.append((row[1].value, normalise(row[2].value)))
    return value_list


def print_attributes(attributes, index=0):
    """Print the collected attributes in a nested structure."""
    attribute, values = attributes[index]
    if len(values) == 1 and values[0][0] == 'X':
        print('%sA-Z' % (' ' * (index + 2)))
    else:
        for char, value in sorted(values):
            print('%s%s v="%s"' % (' ' * (index + 2), char, value))
        print('%sA-Z a="%s"' % (
            ' ' * (index + 2), attribute))
    if index < 3:
        print_attributes(attributes, index + 1)


if __name__ == '__main__':
    # Download the page that contains the link to the current XLS file
    response = requests.get(download_url, timeout=30)
    response.raise_for_status()
    # Find the download link
    document = lxml.html.document_fromstring(response.content)
    links = [a.get('href') for a in document.findall('.//a[@href]')]
    link_url = next(a for a in links if re.match(r'.*/cfi/.*xlsx?$', a))
    # Download and parse the spreadsheet
    response = requests.get(link_url, timeout=30)
    response.raise_for_status()
    workbook = openpyxl.load_workbook(io.BytesIO(response.content), read_only=True)

    print('# generated from %s, downloaded from' % link_url.split('/')[-1])
    print('# %s' % download_url)

    groups = sorted(x for x in workbook.sheetnames if len(x) == 6 and x.endswith('XXXX'))
    for category, name in sorted(get_categories(workbook['Categories'])):
        print('%s category="%s"' % (category, name))
        for group in (x for x in groups if x.startswith(category)):
            sheet = workbook[group]
            print(' %s group="%s"' % (group[1], normalise(sheet.cell(1, 1).value)))
            print_attributes(get_attributes(sheet))
