#!/usr/bin/env python3
# coding: utf-8

# update/be_banks.py - script to download Bank list from Belgian National Bank
#
# Copyright (C) 2018-2025 Arthur de Jong
#
# This library is free software; you can redistribute it and/or
# modify it under the terms of the GNU Lesser General Public
# License as published by the Free Software Foundation; either
# version 2.1 of the License, or (at your option) any later version.
#
# This library is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the GNU
# Lesser General Public License for more details.
#
# You should have received a copy of the GNU Lesser General Public
# License along with this library; if not, write to the Free Software
# Foundation, Inc., 51 Franklin Street, Fifth Floor, Boston, MA
# 02110-1301 USA

"""This script downloads the list of banks with bank codes as used in the
IBAN and BIC codes as published by the Belgian National Bank."""

import io
import os.path

import openpyxl
import requests


# The location of the XLS version of the bank identification codes. Also see
# https://www.nbb.be/en/payment-systems/payment-standards/bank-identification-codes
download_url = 'https://www.nbb.be/doc/be/be/protocol/current_codes.xlsx'


# List of values that refer to non-existing, reserved or otherwise not-
# allocated entries.
not_applicable_values = (
    '-',
    'Indisponible',
    'LIBRE',
    'NAP',
    'NAV',
    'NYA',
    'Onbeschikbaar',
    'VRIJ - LIBRE',
    'VRIJ',
    'nav',
)


def clean(value):
    """Clean up and convert read values removing various placeholder names."""
    value = value.strip()
    if value not in not_applicable_values:
        return value
    return ''


def get_values(sheet):
    """Return values (from, to, bic, bank_name) from the worksheet."""
    rows = sheet.iter_rows()
    # skip first two rows
    try:
        next(rows)
        next(rows)
    except StopIteration:
        pass  # ignore empty worksheet
    # go over rows with values
    for row in rows:
        row = [clean(column.value) for column in row]
        low, high, bic = row[:3]
        bank = ([x for x in row[3:] if x] + [''])[0]
        if bic or bank:
            yield low, high, bic.replace(' ', ''), bank


if __name__ == '__main__':
    response = requests.get(download_url, timeout=30)
    response.raise_for_status()
    workbook = openpyxl.load_workbook(io.BytesIO(response.content), read_only=True)
    sheet = workbook.worksheets[0]
    version = sheet.cell(1, 1).value
    print('# generated from %s downloaded from' %
          os.path.basename(download_url))
    print('# %s' % download_url)
    print('# %s' % version)
    for low, high, bic, bank in get_values(sheet):
        info = '%s-%s' % (low, high)
        if bic:
            info += ' bic="%s"' % bic
        if bank:
            info += ' bank="%s"' % bank
        print(info)
